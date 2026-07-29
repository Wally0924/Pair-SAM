"""
產出 ACDC val/test per-class IoU 比較表 Excel。
  Sheet 1: per-class IoU (CMA / Refign-DAFormer / ours)
  Sheet 2: per-condition mIoU (ours, 4 conditions)
  Sheet 3: 註解（資料集口徑差異）
"""
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


_THIS = Path(__file__).resolve()
RESULTS_JSON = _THIS.parents[3] / 'docs' / 'experiments' / 'v15-eval-2026-05-14' / 'e1_acdc_val_paper_results.json'
OUT_XLSX = _THIS.parents[3] / 'docs' / 'experiments' / 'v15-eval-2026-05-14' / 'comparison_v15_E27.xlsx'

# Cityscapes 19 classes（與論文 Table 順序對齊）
CLASS_KEYS = [
    'road', 'sidewalk', 'building', 'wall', 'fence', 'pole',
    'traffic light', 'traffic sign', 'vegetation', 'terrain', 'sky',
    'person', 'rider', 'car', 'truck', 'bus', 'train', 'motorcycle', 'bicycle',
]
CLASS_DISPLAY = [
    'road', 'sidew.', 'build.', 'wall', 'fence', 'pole',
    'light', 'sign', 'veget.', 'terrain', 'sky',
    'person', 'rider', 'car', 'truck', 'bus', 'train', 'motorc.', 'bicycle',
]

# 來自 CMA paper Table 1（SegFormer, ACDC test）
CMA_TEST = [94.0, 75.2, 88.6, 50.5, 45.5, 54.9, 65.7, 64.2, 87.1,
            61.3, 95.2, 67.0, 45.2, 86.2, 68.6, 76.6, 83.9, 43.3, 60.5]
CMA_TEST_MEAN = 69.1

# 來自 Refign paper Table 1（DAFormer, ACDC test）
REFIGN_TEST = [89.5, 63.4, 87.3, 43.6, 34.3, 52.3, 63.2, 61.4, 86.9,
               58.5, 95.7, 62.1, 39.3, 84.1, 65.7, 71.3, 85.4, 47.9, 52.8]
REFIGN_TEST_MEAN = 65.5

# 論文 ablation 給的 val 數字（沒有 per-class，只有 overall）
CMA_VAL_MEAN = 67.2     # CMA Table 5 row 7
REFIGN_VAL_MEAN = 65.0  # Refign Table 4 row 6


def main():
    data = json.loads(RESULTS_JSON.read_text())
    ours = [data['per_class_iou_overall'][c] * 100 for c in CLASS_KEYS]
    ours_mean = data['overall_miou'] * 100
    per_cond = {k: v * 100 for k, v in data['per_condition_miou'].items()}

    wb = openpyxl.Workbook()

    # ── Sheet 1: per-class IoU ──
    ws = wb.active
    ws.title = 'Per-Class IoU'

    header_fill = PatternFill('solid', fgColor='4472C4')
    header_font = Font(bold=True, color='FFFFFF')
    ours_fill = PatternFill('solid', fgColor='FFF2CC')
    thin = Side(border_style='thin', color='808080')
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')

    # Header
    headers = ['Method', 'Eval Split'] + CLASS_DISPLAY + ['mean']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = box

    rows = [
        ('CMA (SegFormer)',         'ACDC test (2000)', CMA_TEST,    CMA_TEST_MEAN),
        ('Refign-DAFormer',         'ACDC test (2000)', REFIGN_TEST, REFIGN_TEST_MEAN),
        ('Ours (PairSAM v15 E27)', 'ACDC val (406)',   ours,       ours_mean),
    ]

    for r_idx, (name, split, vals, mean) in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=name).border = box
        ws.cell(row=r_idx, column=2, value=split).border = box
        for ci, v in enumerate(vals, start=3):
            cell = ws.cell(row=r_idx, column=ci, value=round(v, 2))
            cell.number_format = '0.00'
            cell.alignment = center
            cell.border = box
        mean_cell = ws.cell(row=r_idx, column=3 + len(vals), value=round(mean, 2))
        mean_cell.font = Font(bold=True)
        mean_cell.number_format = '0.00'
        mean_cell.alignment = center
        mean_cell.border = box

        if name.startswith('Ours'):
            for c in range(1, 3 + len(vals) + 1):
                ws.cell(row=r_idx, column=c).fill = ours_fill

    # 欄寬
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 18
    for c in range(3, 3 + len(CLASS_DISPLAY) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 9

    # 第二區塊：val-to-val overall（公平比較）
    ws.cell(row=6, column=1, value='Fair val-to-val comparison (overall mIoU only)').font = Font(bold=True, italic=True)
    fair_headers = ['Method', 'ACDC val mIoU (%)']
    for c, h in enumerate(fair_headers, 1):
        cell = ws.cell(row=7, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = box

    fair_rows = [
        ('CMA (SegFormer, Table 5 row 7)', CMA_VAL_MEAN),
        ('Refign-DAFormer (Table 4 row 6)', REFIGN_VAL_MEAN),
        ('Ours (PairSAM v15 E27)',       ours_mean),
    ]
    for r_idx, (name, val) in enumerate(fair_rows, start=8):
        c1 = ws.cell(row=r_idx, column=1, value=name)
        c2 = ws.cell(row=r_idx, column=2, value=round(val, 2))
        c1.border = box; c2.border = box
        c2.alignment = center; c2.number_format = '0.00'
        if name.startswith('Ours'):
            c1.fill = ours_fill; c2.fill = ours_fill
            c1.font = Font(bold=True); c2.font = Font(bold=True)

    # ── Sheet 2: per-condition mIoU ──
    ws2 = wb.create_sheet('Per-Condition mIoU (Ours)')
    ws2['A1'] = 'Condition'; ws2['B1'] = 'mIoU (%)'; ws2['C1'] = 'Samples'
    for c in 'ABC':
        ws2[f'{c}1'].fill = header_fill
        ws2[f'{c}1'].font = header_font
        ws2[f'{c}1'].alignment = center
        ws2[f'{c}1'].border = box

    cond_rows = [
        ('Fog',   per_cond['fog'],   100),
        ('Rain',  per_cond['rain'],  100),
        ('Snow',  per_cond['snow'],  100),
        ('Night', per_cond['night'], 106),
        ('All',   ours_mean,         406),
    ]
    for r_idx, (cond, miou, n) in enumerate(cond_rows, start=2):
        ws2.cell(row=r_idx, column=1, value=cond).border = box
        m = ws2.cell(row=r_idx, column=2, value=round(miou, 2))
        m.number_format = '0.00'; m.alignment = center; m.border = box
        ws2.cell(row=r_idx, column=3, value=n).alignment = center
        ws2.cell(row=r_idx, column=3).border = box
        if cond == 'All':
            for c in 'ABC':
                ws2[f'{c}{r_idx}'].font = Font(bold=True)
                ws2[f'{c}{r_idx}'].fill = ours_fill

    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 12

    # ── Sheet 3: Notes ──
    ws3 = wb.create_sheet('Notes')
    notes = [
        ('Notes on evaluation protocols', ''),
        ('', ''),
        ('Source — CMA',       'Bruggemann et al., ICCV 2023. Table 1 (SegFormer, ACDC test 2000).'),
        ('Source — Refign',    'Bruggemann et al., WACV 2023. Table 1 (DAFormer, ACDC test 2000).'),
        ('Source — Ours',      'PairSAM v15, checkpoint best_E27_mIoU65.68_LR4.0e-05.pth, ACDC val 406.'),
        ('', ''),
        ('Resolution',         'Ours: native 1080x1920 (pred upsampled bilinear from 256x256 logits to 1080x1920, GT not resized).'),
        ('Train/eval protocol','Ours: 1024x1024 input to ViT, then upsample. Confusion-matrix based mIoU.'),
        ('Ignore index',       '255 (GT) and ACDC invalid_mask combined as ignored.'),
        ('', ''),
        ('val vs test',        'CMA / Refign per-class numbers are from ACDC test set (server eval).'),
        ('',                   'Ours is on ACDC val (406 images). For fair val-to-val see Sheet 1 lower block.'),
        ('',                   'CMA val mIoU = 67.2% (Table 5 row 7), Refign val mIoU = 65.0% (Table 4 row 6).'),
        ('',                   'Per-class val numbers are not provided in either paper.'),
        ('', ''),
        ('Generated',          'export_comparison_xlsx.py'),
    ]
    for r_idx, (k, v) in enumerate(notes, start=1):
        ws3.cell(row=r_idx, column=1, value=k).font = Font(bold=True) if r_idx == 1 else Font()
        ws3.cell(row=r_idx, column=2, value=v)
    ws3.column_dimensions['A'].width = 22
    ws3.column_dimensions['B'].width = 100

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f'✅ Excel written: {OUT_XLSX}')


if __name__ == '__main__':
    main()
